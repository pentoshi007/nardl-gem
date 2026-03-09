"""
IIP Chain-Linking Script
========================
Reads IIP.xlsx (downloaded from RBI DBIE), chain-links the
Base 2004-05 and Base 2011-12 General Index series using the
overlap-ratio method, and outputs a clean iip_chained.xlsx.

HOW TO RUN:
  1. Place this script in the same folder as IIP.xlsx
  2. Open terminal / command prompt in that folder
  3. Run:  python chain_link_iip.py
  4. Output file:  iip_chained.xlsx  (same folder)

REQUIREMENTS:
  pip install openpyxl pandas matplotlib
"""

import re
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from datetime import date

# ── 1. Load the sheet ────────────────────────────────────────────────────────
print("Reading IIP.xlsx ...")
wb = openpyxl.load_workbook("IIP.xlsx")

sheet_name = next(s for s in wb.sheetnames if "Sector" in s)
ws = wb[sheet_name]
rows = [list(r) for r in ws.iter_rows(values_only=True)]
print(f"  Sheet found: '{sheet_name}'")

# ── 2. Helper: parse "2012:04 (APR)" → date(2012,4,1) ───────────────────────
def parse_date(s):
    if not s or not isinstance(s, str):
        return None
    m = re.match(r"(\d{4}):(\d{2})", s.strip())
    if m:
        return date(int(m.group(1)), int(m.group(2)), 1)
    return None

# ── 3. Extract Base 2011-12 (dates row 6, General Index row 9) ──────────────
print("\nExtracting Base 2011-12 General Index ...")
dates_11_raw  = [x for x in rows[5][2:] if x is not None]
values_11_raw = [x for x in rows[8][2:] if x is not None]

dates_11  = [parse_date(d) for d in reversed(dates_11_raw)]
values_11 = list(reversed(values_11_raw))

df_11 = pd.DataFrame({"date": dates_11, "iip": values_11}).dropna()
df_11["date"] = pd.to_datetime(df_11["date"])
df_11 = df_11.sort_values("date").reset_index(drop=True)
print(f"  Base 2011-12: {df_11['date'].min().strftime('%b %Y')} -> {df_11['date'].max().strftime('%b %Y')}  ({len(df_11)} months)")

# ── 4. Extract Base 2004-05 (dates row 14, General Index row 18) ────────────
print("\nExtracting Base 2004-05 General Index ...")
dates_04_raw  = [x for x in rows[13][2:] if x is not None]
values_04_raw = [x for x in rows[17][2:] if x is not None]

dates_04  = [parse_date(d) for d in reversed(dates_04_raw)]
values_04 = list(reversed(values_04_raw))

df_04 = pd.DataFrame({"date": dates_04, "iip": values_04}).dropna()
df_04["date"] = pd.to_datetime(df_04["date"])
df_04 = df_04.sort_values("date").reset_index(drop=True)
print(f"  Base 2004-05: {df_04['date'].min().strftime('%b %Y')} -> {df_04['date'].max().strftime('%b %Y')}  ({len(df_04)} months)")

# ── 5. Compute splice factor over overlap period ─────────────────────────────
overlap_start = pd.Timestamp("2012-04-01")
overlap_end   = pd.Timestamp("2017-01-01")

overlap_11 = df_11[(df_11["date"] >= overlap_start) & (df_11["date"] <= overlap_end)]["iip"]
overlap_04 = df_04[(df_04["date"] >= overlap_start) & (df_04["date"] <= overlap_end)]["iip"]

print(f"\nOverlap period: {overlap_start.strftime('%b %Y')} - {overlap_end.strftime('%b %Y')}")
print(f"  Overlap months in 2011-12 series: {len(overlap_11)}")
print(f"  Overlap months in 2004-05 series: {len(overlap_04)}")

splice_factor = overlap_11.mean() / overlap_04.mean()
print(f"\n  Splice factor = {overlap_11.mean():.4f} / {overlap_04.mean():.4f} = {splice_factor:.6f}")
print(f"  (Expected approx 0.627  --  if very different, stop and check the file)")

# ── 6. Apply splice and concatenate ──────────────────────────────────────────
df_04["iip_spliced"] = df_04["iip"] * splice_factor

cutoff = pd.Timestamp("2012-04-01")
part_old = df_04[df_04["date"] < cutoff][["date","iip_spliced"]].rename(columns={"iip_spliced":"iip_chained"})
part_new = df_11[df_11["date"] >= cutoff][["date","iip"]].rename(columns={"iip":"iip_chained"})

df_full = pd.concat([part_old, part_new]).sort_values("date").reset_index(drop=True)

# ── 7. Filter to study window: April 2004 to December 2024 ──────────────────
start = pd.Timestamp("2004-04-01")
end   = pd.Timestamp("2024-12-01")
df_final = df_full[(df_full["date"] >= start) & (df_full["date"] <= end)].reset_index(drop=True)

actual_end = df_final["date"].max()
n_obs = len(df_final)

print(f"\n  Final series: {df_final['date'].min().strftime('%b %Y')} -> {actual_end.strftime('%b %Y')}")
print(f"  Total observations: {n_obs}")

if n_obs == 249:
    print(f"  OK - 249 observations as expected (Apr 2004 - Dec 2024)")
else:
    print(f"  WARNING: Got {n_obs} observations (expected 249).")
    print(f"  IIP data in your file ends at {actual_end.strftime('%b %Y')}.")
    print("  Check source file coverage and study window filters.")

# ── 8. Plot to verify no discontinuity ──────────────────────────────────────
print("\nGenerating verification plot ...")
fig, ax = plt.subplots(figsize=(14, 5))
ax.plot(df_final["date"], df_final["iip_chained"],
        color="#1F3864", linewidth=1.2, label="IIP Chained (General Index)")
ax.axvline(x=cutoff, color="red", linestyle="--", linewidth=1.2, label="Splice point (Apr 2012)")
ax.set_title(
    f"India IIP General Index - Chain-Linked\n"
    f"Apr 2004 to {actual_end.strftime('%b %Y')}  |  Splice factor = {splice_factor:.4f}"
)
ax.set_xlabel("Date")
ax.set_ylabel("IIP Index (Base 2011-12 equivalent units)")
ax.legend()
ax.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig("iip_verification_plot.png", dpi=150)
print("  Saved: iip_verification_plot.png")
print("  CHECK: There must be NO visible jump at the red dashed line.")

# ── 9. Save to Excel ─────────────────────────────────────────────────────────
df_out = df_final.copy()
df_out["date"] = df_out["date"].dt.strftime("%Y-%m-%d")
df_out = df_out[["date", "iip_chained"]]

with pd.ExcelWriter("iip_chained.xlsx", engine="openpyxl") as writer:
    df_out.to_excel(writer, index=False, sheet_name="IIP_Chained")

    summary = pd.DataFrame({
        "Parameter": [
            "Source file",
            "Sheet used",
            "Base 2004-05  dates row / values row",
            "Base 2011-12  dates row / values row",
            "Overlap period",
            "Overlap months",
            "Splice factor",
            "Study start",
            "Study end (actual in this file)",
            "Total observations",
            "Column: date",
            "Column: iip_chained",
        ],
        "Value": [
            "IIP.xlsx (RBI DBIE)",
            sheet_name,
            "Row 14 / Row 18",
            "Row 6 / Row 9",
            "Apr 2012 - Jan 2017",
            str(len(overlap_11)),
            f"{splice_factor:.6f}",
            "Apr 2004",
            actual_end.strftime("%b %Y"),
            str(n_obs),
            "Date as YYYY-MM-DD (first of each month)",
            "IIP General Index in Base 2011-12 equivalent units",
        ]
    })
    summary.to_excel(writer, index=False, sheet_name="Metadata")

print(f"\n  Saved: iip_chained.xlsx")
print(f"  Sheet IIP_Chained : {n_obs} rows  |  Apr 2004 - {actual_end.strftime('%b %Y')}")
print(f"  Sheet Metadata    : splice parameters for your records")
print()
print("=" * 60)
print("NEXT STEP: Download 3 series from FRED")
print("=" * 60)
print()
print("  Go to https://fred.stlouisfed.org and download each as CSV.")
print("  Do NOT filter on the website. Download full series and trim in R.")
print()
print("  Series to download:")
print()
print("  1. POILBREUSDM  (Brent Crude, USD/barrel, monthly)")
print("     Download from: Jan 2000  |  to: latest available")
print()
print("  2. INDCPIALLMINMEI  (India CPI, OECD, monthly index)")
print("     Download from: Jan 2000  |  to: latest available")
print("     NOTE: This series has a publication lag.")
print("     If it ends before Dec 2024, your study end date is Dec 2024 anyway.")
print("     If it ends at or after Dec 2024, trim to Dec 2024 in R.")
print()
print("  3. EXINUS  (INR/USD exchange rate, monthly)")
print("     Download from: Jan 2000  |  to: latest available")
print("     IMPORTANT: use EXINUS (monthly), NOT DEXINUS (daily)")
print()
print("  Study window in R: April 2004 - December 2024  (N = 249)")
print("  Pre-sample runway: Jan 2000 - Mar 2004 gives R enough data")
print("  to compute up to 12 lags without eating into your study window.")
